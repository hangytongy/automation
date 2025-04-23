import smtplib
from email.message import EmailMessage
from datetime import datetime, timedelta
import openpyxl
import os
import win32com.client
from openpyxl.drawing.image import Image
from PIL import Image as PILImage


def resize_image_to_fit_cell(image_path, target_width_px, target_height_px):
    img = PILImage.open(image_path)
    img = img.resize((target_width_px, target_height_px), PILImage.LANCZOS)
    resized_path = "resized_" + image_path
    img.save(resized_path)
    return resized_path

column_width = 30  # in Excel units
row_height = 16    # in points

# Approximate conversion to pixels
target_width_px = int(column_width * 7.5)
target_height_px = int(row_height * 1.33)

# Resize the image
image_path = "logo_secondary_fullblack.png"
resized_path = resize_image_to_fit_cell(image_path, target_width_px, target_height_px)

def get_invoice_no(project):
    date = datetime.now()
    date_obj = datetime.strptime(project["start date"], "%d/%m/%Y")

    time_delta = date - date_obj
    no_of_invoices = int((time_delta/30).days)
    invoice_no = project["initail inv no"] + no_of_invoices - 1
    
    return invoice_no,no_of_invoices

def get_start_end_date(project,no_of_invoices):
    
    date_obj = datetime.strptime(project["start date"], "%d/%m/%Y")
    start_date = date_obj + timedelta(days = 30 * no_of_invoices)
    end_date = start_date + timedelta(days = 30) - timedelta(seconds = 1)
    
    start_date = start_date.date().strftime("%d %b'%y")
    end_date = end_date.date().strftime("%d %b'%y")
    
    return start_date, end_date

#make the discription in the df into dictionary
def get_invoice_items(project,start_date,end_date):
    
    date_range = f"{start_date} - {end_date}"
    
    invoice_items = [
        {
            "discription" : f"{desc} \n {date_range}" ,
             "amount" : amt
        } 
        for desc,amt in zip(project["discription"], project["amount"])
    ]
    
    return invoice_items

def fill_excel_template(template_path, img, client_name, client_alias ,client_address, invoice_no, 
                        date, inv_date, invoice_items, note, payment_option,project_folder):
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Fill invoice content 
    
    #Add MF logo
    ws.add_image(img, 'C3')
    
    #client name
    ws['C10'] = client_name
    
    #client address
    ws['C11'] = client_address
    
    #date of invoice
    ws['F10'] = date
    
    #invoice num
    ws['F7'] = f'#Inv-{inv_date}-{client_alias}-{invoice_no}'

    row = 16
    total = 0
    for item in invoice_items:
        ws[f"C{row}"] = item['discription']
        ws[f"E{row}"] = item['amount']
        total += item['amount']
        row += 1
        
    #Note
    ws['C27'] = note['teams']
    ws['C28'] = note['start']
    ws['C29'] = note['end']
    
    #Payment option
    ws['C32'] = f"{total:,} {payment_option['payment']}"
    ws['C33'] = payment_option['address']
    

    output_excel_path = f'Invoice Inv-{inv_date}-{client_alias}-{invoice_no}.xlsx'

    output_excel_path = os.path.join(project_folder,output_excel_path)
    
    wb.save(output_excel_path)
    print(f"✅ Excel invoice saved at {output_excel_path}")
    return output_excel_path

def convert_excel_to_pdf(excel_path):

    output_pdf_path = f"{excel_path.split('.')[0]}.pdf"

    excel = win32com.client.Dispatch("Excel.Application")
    wb = excel.Workbooks.Open(os.path.abspath(excel_path))
    wb.ExportAsFixedFormat(0, os.path.abspath(output_pdf_path))
    wb.Close()
    excel.Quit()

    os.remove(excel_path)

    print(f"✅ PDF exported to {output_pdf_path}")
    return output_pdf_path

def create_invoice(project_folder,client_name,client_addy,client_alias,invoice_no,invoice_items,no_of_teams,start_date,
                   end_date,start_time,payment_type,payment_addy):
    

    start_time_obj = datetime.strptime(start_time, "%H%M")
    end_time_obj = start_time_obj - timedelta(seconds = 1)
    end_time = end_time_obj.strftime("%H%M")
    
    img = Image("resized_logo_secondary_fullblack.png")
    client = client_name
    client_alias = client_alias
    client_address = client_addy
    invoice_no = invoice_no
    date = datetime.now().strftime("%B %d, %Y")
    inv_date = datetime.now().strftime("%m%d")
    invoice_items = invoice_items
    note = {
        "teams" : f"{no_of_teams} team of 3x Community Managers will be provided for Discord & Intercom for the following period:",
        "start" : f"Start date: {start_date}, {start_time} GMT+8",
        "end" : f"End date: {end_date}, {end_time} GMT+8"
    } 
    payment_option = {
        "payment" : f"USDC to be made to the {payment_type} address:",
        "address" : f"{payment_addy}"
            }

    excel_path = fill_excel_template("standard_invoice.xlsx", img, client,client_alias, client_address,
                                     invoice_no, date, inv_date,invoice_items, note, payment_option,project_folder)
    output_pdf_path = convert_excel_to_pdf(excel_path)

    return output_pdf_path


